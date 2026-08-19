import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 1. 원본 데이터 불러오기
raw_file = "물류센터 매입매출현황(품목별)_20260819160343.xlsx"
df = pd.read_excel(raw_file)

# 2. 숫자 데이터 전처리 (문자열 내 콤마 제거 및 숫자로 변환)
num_cols = ['입고', '매출단가', '매출금', '이익금']
for col in num_cols:
    df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)

# 3. 유의미한 데이터 필터링 (매출금이 발생한 내역만 복사)
df_filtered = df[df['매출금'] > 0].copy()

# 4. 핵심 로직: 제조사 분리 및 값 1/n 연산
# 엔터('\n')를 기준으로 제조사 문자열을 리스트로 쪼개기
df_filtered['제조사_list'] = df_filtered['제조사'].astype(str).str.split('\n')

# 각 행마다 묶여있는 제조사의 개수(n) 파악
df_filtered['n'] = df_filtered['제조사_list'].apply(len)

# 1/n로 수량(입고), 매출단가, 매출금, 이익금 나누기
for col in num_cols:
    df_filtered[col] = df_filtered[col] / df_filtered['n']

# 리스트 내에 있던 제조사들을 각각 새로운 독립된 행으로 펄치기(Explode)
df_exploded = df_filtered.explode('제조사_list')
df_exploded['제조사'] = df_exploded['제조사_list'].str.strip() # 이름 앞뒤 공백 제거

# 5. 제조사, 품목, 규격별로 그룹화 및 합산
report_df = df_exploded.groupby(['제조사', '품목', '규격']).agg({
    '입고': 'sum',
    '매출단가': 'sum',
    '매출금': 'sum',
    '이익금': 'sum'
}).reset_index()

# 6. 보기 좋게 정렬 및 엑셀용 컬럼명 변경
report_df.sort_values(['제조사', '품목'], inplace=True)
report_df.rename(columns={
    '제조사': '제조사(매입처)',
    '품목': '품목명',
    '규격': '규격',
    '입고': '수량',
    '매출단가': '매출단가(원)',
    '매출금': '매출금(원)',
    '이익금': '이익금(원)'
}, inplace=True)

# 7. 엑셀 파일 생성 (openpyxl을 이용해 셀 스타일 지정)
output_file = "물류수익보고서_제조사분리_완료.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "수익보고서"

# 데이터프레임을 엑셀 워크시트에 삽입
for r in dataframe_to_rows(report_df, index=False, header=True):
    ws.append(r)

# --- 엑셀 디자인 (표 테두리 및 헤더 색상) ---
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin_border = Border(
    left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), 
    top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF")
)

# 1행(헤더) 스타일 지정
for col in range(1, len(report_df.columns) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 본문 내용 스타일 및 숫자 콤마 천 단위 포맷 적용
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if col in [1, 2, 3]: # 텍스트 컬럼 (제조사, 품목, 규격)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        if col in [4, 5, 6, 7]: # 숫자 컬럼 (수량, 단가, 매출금, 이익금)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '#,##0' # 천 단위 콤마 포맷 (소수점은 반올림되어 보임)

# 셀 너비 최적화
ws.column_dimensions['A'].width = 30 # 제조사
ws.column_dimensions['B'].width = 40 # 품목
ws.column_dimensions['C'].width = 15 # 규격
ws.column_dimensions['D'].width = 10 # 수량
ws.column_dimensions['E'].width = 15 # 매출단가
ws.column_dimensions['F'].width = 20 # 매출금
ws.column_dimensions['G'].width = 20 # 이익금

# 최종 저장
wb.save(output_file)